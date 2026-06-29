import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Load JSONs
with open("translations/ua.json", "r", encoding="utf-8") as f:
    ua = json.load(f)
with open("translations/eu.json", "r", encoding="utf-8") as f:
    eu = json.load(f)
with open("translations/pl.json", "r", encoding="utf-8") as f:
    pl = json.load(f)

# Group keys logically
def get_group(k):
    if k.startswith("home.hero"):
        return "Home - Hero Section"
    elif k.startswith("home.problem"):
        return "Home - Problem Section"
    elif k.startswith("home.solution"):
        return "Home - Solution Section"
    elif k.startswith("home.calc"):
        return "Home - Calculations CTA"
    elif k.startswith("home.trust"):
        return "Home - Trust Section"
    elif k.startswith("home.team"):
        return "Home - Team Section"
    elif k.startswith("home.grants"):
        return "Home - Grants Section"
    elif k.startswith("tech.hero"):
        return "Technology - Hero Section"
    elif k.startswith("tech.htc"):
        return "Technology - HTC Concept"
    elif k.startswith("tech.configs"):
        return "Technology - Configurations"
    elif k.startswith("tech.hydrochar"):
        return "Technology - Hydrochar Specs"
    elif k.startswith("tech.contrast"):
        return "Technology - Comparison Table"
    elif k.startswith("tech.models"):
        return "Technology - Models Line"
    elif k.startswith("sim"):
        return "Simulator Page"
    elif k.startswith("htc"):
        return "Simulator - Chemical Passport"
    elif k.startswith("chat"):
        return "AI Consultant Chat"
    elif k.startswith("contact") or k.startswith("form"):
        return "Contact & Lead Form"
    elif k.startswith("header") or k.startswith("footer"):
        return "Header, Navigation & Footer"
    else:
        return "General / UI Elements"

# Initialize workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Translations Review"

# Style templates
font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
fill_header = PatternFill(start_color="046BD2", end_color="046BD2", fill_type="solid")
alignment_header = Alignment(horizontal="center", vertical="center", wrap_text=True)

font_data = Font(name="Segoe UI", size=10)
alignment_data = Alignment(horizontal="left", vertical="top", wrap_text=True)
alignment_key = Alignment(horizontal="left", vertical="top")

thin_border = Border(
    left=Side(style='thin', color='DDDDDD'),
    right=Side(style='thin', color='DDDDDD'),
    top=Side(style='thin', color='DDDDDD'),
    bottom=Side(style='thin', color='DDDDDD')
)

# Header Row
headers = ["Key", "Section", "Ukrainian (UA)", "English (EN)", "Polish (PL)", "Proposed Change"]
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = alignment_header
    cell.border = thin_border

# Populate Data
row_idx = 2
for k in sorted(ua.keys()):
    if k.startswith("_comment"):
        continue
    
    val_ua = ua.get(k, "")
    val_en = eu.get(k, "")
    val_pl = pl.get(k, "")
    grp = get_group(k)
    
    vals = [k, grp, val_ua, val_en, val_pl, ""]
    for col_idx, val in enumerate(vals, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = font_data
        cell.alignment = alignment_data if col_idx > 2 else alignment_key
        cell.border = thin_border
        
        # Style proposed change column slightly differently
        if col_idx == 6:
            cell.fill = PatternFill(start_color="FCFDFD", end_color="FCFDFD", fill_type="solid")
            
    row_idx += 1

# Auto-fit column widths with padding
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 50
ws.column_dimensions['D'].width = 50
ws.column_dimensions['E'].width = 50
ws.column_dimensions['F'].width = 50

# Freeze panes so header is always visible
ws.freeze_panes = "A2"

# Enable filters
ws.auto_filter.ref = f"A1:F{row_idx-1}"

# Save
wb.save("translations_review.xlsx")
print("Excel spreadsheet generated successfully.")
